#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
钢铁行业减碳模型分析脚本
"""

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

# 读取Excel文件
file_path = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

print("=" * 80)
print("钢铁行业减碳模型V11分析报告")
print("=" * 80)

# 1. 列出所有Sheet
print("\n1. Sheet列表:")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"   {idx}. {sheet_name}")

# 2. 检查三个情景Sheet的Step 8输出
scenarios = {
    '积极情景': None,
    '适度情景': None,
    '保守情景': None
}

for sheet_name in scenarios.keys():
    if sheet_name in wb.sheetnames:
        scenarios[sheet_name] = wb[sheet_name]

# 3. 读取关键年份的碳强度数据
key_years = [2024, 2025, 2030, 2035, 2040, 2050, 2060]
print("\n2. 三个情景Sheet的Step 8碳强度输出:")

for scenario_name, ws in scenarios.items():
    if ws is None:
        continue
    
    print(f"\n   【{scenario_name}】")
    
    # 找到年份行(Row 4)和碳强度行
    # 根据文档: Row 4是年份, Step 8碳强度在Row 62左右
    year_row = 4
    
    # 查找"吨钢碳排放强度"行
    ci_row = None
    for row in range(39, 66):  # Step 8区域
        cell_value = ws.cell(row, 2).value  # B列
        if cell_value and "吨钢碳排放强度" in str(cell_value):
            ci_row = row
            break
    
    if ci_row:
        print(f"   碳强度行: Row {ci_row}")
        # 读取关键年份数据
        for year in key_years:
            # 找到对应年份的列
            for col in range(3, 42):  # C列到AM列
                year_val = ws.cell(year_row, col).value
                if year_val == year:
                    ci_val = ws.cell(ci_row, col).value
                    if isinstance(ci_val, str) and '=' in ci_val:
                        print(f"   {year}: 公式存在 (需要计算)")
                    else:
                        print(f"   {year}: {ci_val}")
                    break
    else:
        print("   未找到碳强度行")

# 4. 读取Dashboard目标值
dashboard_sheet_name = "汇总对比与可视化看板"
if dashboard_sheet_name in wb.sheetnames:
    ws_dashboard = wb[dashboard_sheet_name]
    print(f"\n3. Dashboard目标值 ({dashboard_sheet_name}):")
    
    # 查找碳强度目标数据区域
    # 根据文档: B列(保守)、D列(适度)、F列(积极)
    print("\n   关键年份碳强度目标:")
    print("   年份 | 积极 | 适度 | 保守")
    print("   " + "-" * 40)
    
    for row in range(5, 50):  # 搜索数据区域
        year_val = ws_dashboard.cell(row, 1).value  # A列年份
        if year_val in key_years:
            aggressive = ws_dashboard.cell(row, 6).value  # F列积极
            moderate = ws_dashboard.cell(row, 4).value    # D列适度
            conservative = ws_dashboard.cell(row, 2).value  # B列保守
            print(f"   {year_val} | {aggressive} | {moderate} | {conservative}")

# 5. 检查Step 3产量数据
print("\n4. Step 3 粗钢产量检查:")
for scenario_name, ws in scenarios.items():
    if ws is None:
        continue
    
    # 查找"粗钢产量"行 (Row 12)
    prod_row = 12
    for row in range(10, 15):
        cell_value = ws.cell(row, 2).value
        if cell_value and "粗钢产量" in str(cell_value):
            prod_row = row
            break
    
    print(f"\n   【{scenario_name}】粗钢产量 (Row {prod_row}):")
    for year in [2024, 2030, 2035, 2050, 2060]:
        for col in range(3, 42):
            year_val = ws.cell(4, col).value
            if year_val == year:
                prod_val = ws.cell(prod_row, col).value
                print(f"   {year}: {prod_val} 万吨")
                break

wb.close()

print("\n" + "=" * 80)
print("分析完成")
print("=" * 80)
