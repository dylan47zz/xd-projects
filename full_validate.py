#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""完整验证V11.1所有年份对齐情况及关键指标"""
import openpyxl

wb = openpyxl.load_workbook('/Users/daiyinlong/Desktop/钢铁行业模型V11.1.xlsx', data_only=True)

YEARS = list(range(2024, 2061))
YC = 3
BASE_CI = 1.95

dashboard = {
    '积极情景': {2024:1.9500,2025:1.9470,2026:1.9050,2027:1.8530,2028:1.8120,2029:1.7800,
                 2030:1.7440,2031:1.6690,2032:1.5940,2033:1.5200,2034:1.4450,2035:1.3700,
                 2036:1.2660,2037:1.1620,2038:1.0580,2039:0.9540,2040:0.8500,2041:0.7820,
                 2042:0.7140,2043:0.6460,2044:0.5780,2045:0.5100,2046:0.4420,2047:0.3740,
                 2048:0.3060,2049:0.2380,2050:0.1700,2051:0.1540,2052:0.1380,2053:0.1220,
                 2054:0.1060,2055:0.0900,2056:0.0740,2057:0.0580,2058:0.0420,2059:0.0260,2060:0.0100},
    '适度情景': {2024:1.9500,2025:1.9540,2026:1.9230,2027:1.8910,2028:1.8600,2029:1.8290,
                 2030:1.7980,2031:1.7600,2032:1.7220,2033:1.6840,2034:1.6460,2035:1.6070,
                 2036:1.5160,2037:1.4240,2038:1.3330,2039:1.2410,2040:1.1500,2041:1.0750,
                 2042:1.0000,2043:0.9250,2044:0.8500,2045:0.7750,2046:0.7000,2047:0.6250,
                 2048:0.5500,2049:0.4750,2050:0.4000,2051:0.3650,2052:0.3300,2053:0.2950,
                 2054:0.2600,2055:0.2250,2056:0.1900,2057:0.1550,2058:0.1200,2059:0.0850,2060:0.0500},
    '保守情景': {2024:1.9500,2025:1.9580,2026:1.9410,2027:1.9240,2028:1.9070,2029:1.8910,
                 2030:1.8740,2031:1.8490,2032:1.8250,2033:1.8000,2034:1.7760,2035:1.7510,
                 2036:1.6910,2037:1.6310,2038:1.5710,2039:1.5100,2040:1.4500,2041:1.3800,
                 2042:1.3100,2043:1.2400,2044:1.1700,2045:1.1000,2046:1.0300,2047:0.9600,
                 2048:0.8900,2049:0.8200,2050:0.7500,2051:0.6870,2052:0.6240,2053:0.5610,
                 2054:0.4980,2055:0.4350,2056:0.3720,2057:0.3090,2058:0.2460,2059:0.1830,2060:0.1200}
}

print("===== 全年份CI对齐验证 =====")
for sc in ['积极情景','适度情景','保守情景']:
    ws = wb[sc]
    print(f"\n【{sc}】所有年份:")
    bad=[]
    for y in YEARS:
        col=YC+(y-2024)
        ci=ws.cell(63,col).value
        P=ws.cell(13,col).value or 0
        tgt=dashboard[sc].get(y,0)
        if ci is not None and tgt:
            dev=(ci-tgt)/tgt*100
            if abs(dev)>1.0: bad.append((y,ci,tgt,dev))
    if bad:
        print(f"  ⚠️ {len(bad)}年偏差>1%:")
        for y,ci,tgt,dev in bad:
            print(f"     {y}: CI={ci:.4f}, 目标={tgt:.4f}, 偏差={dev:.1f}%")
    else:
        print(f"  ✅ 全部37年偏差≤1%")
    # 打印关键数值
    for y in [2025,2030,2035,2040,2050,2060]:
        col=YC+(y-2024)
        ci=ws.cell(63,col).value
        tgt=dashboard[sc].get(y)
        if ci and tgt:
            dev=(ci-tgt)/tgt*100
            print(f"  {y}: CI={ci:.4f}, 目标={tgt:.4f}, 偏差={dev:+.3f}%")

print("\n===== 三情景产量差异验证 =====")
for y in [2024,2025,2030,2035,2040,2050,2060]:
    vals=[wb[sc].cell(13,YC+(y-2024)).value for sc in ['积极情景','适度情景','保守情景']]
    print(f"  {y}: 积极={vals[0]:.0f}, 适度={vals[1]:.0f}, 保守={vals[2]:.0f}")

print("\n===== 2030年短流程占比验证 =====")
col30=YC+6
for sc in ['积极情景','适度情景','保守情景']:
    ws=wb[sc]
    eaf=ws.cell(15,col30).value
    h2=ws.cell(16,col30).value
    print(f"  {sc}: 短流程={eaf:.1f}%, 氢冶金={h2:.2f}%")

print("\n===== CCUS关键指标 =====")
for sc in ['积极情景','适度情景','保守情景']:
    ws=wb[sc]
    P_60=ws.cell(13,YC+36).value or 0
    ccus_r60=ws.cell(59,YC+36).value or 0
    ct60=ws.cell(57,YC+36).value or 0
    ccus_amt60=ct60*ccus_r60
    bau60=P_60*BASE_CI
    net60=ct60*(1-ccus_r60)
    total_reduce=bau60-net60
    contr=ccus_amt60/total_reduce*100 if total_reduce>0 else 0
    P_30=ws.cell(13,YC+6).value or 0
    ccus_r30=ws.cell(59,YC+6).value or 0
    ct30=ws.cell(57,YC+6).value or 0
    ccus_amt30=ct30*ccus_r30
    P_35=ws.cell(13,YC+11).value or 0
    ccus_r35=ws.cell(59,YC+11).value or 0
    ct35=ws.cell(57,YC+11).value or 0
    ccus_amt35=ct35*ccus_r35
    print(f"\n  {sc}:")
    print(f"    2030 CCUS: {ccus_amt30:.0f}万吨={ccus_amt30/10000:.3f}亿吨, 比例={ccus_r30*100:.2f}%")
    print(f"    2035 CCUS: {ccus_amt35:.0f}万吨={ccus_amt35/10000:.3f}亿吨, 比例={ccus_r35*100:.2f}%")
    print(f"    2060 CCUS: {ccus_amt60:.0f}万吨={ccus_amt60/10000:.3f}亿吨, 比例={ccus_r60*100:.2f}%")
    print(f"    2060 贡献率: {contr:.1f}% (prd2要求15-25%，知识文档5.4亿吨天花板)")

print("\n===== 废钢资源量参考数据验证 =====")
for sc in ['积极情景','适度情景','保守情景']:
    ws=wb[sc]
    vals=[(y, ws.cell(22,YC+(y-2024)).value) for y in [2024,2025,2030,2035,2040,2050,2060]]
    print(f"  {sc}: {vals}")

print("\n===== 2030年适度情景验证（prd2最高原则）=====")
ws_mod=wb['适度情景']
ci_2030=ws_mod.cell(63,YC+6).value
ci_2024=ws_mod.cell(63,YC+0).value
if ci_2030 and ci_2024:
    pct_drop=(ci_2024-ci_2030)/ci_2024*100
    print(f"  适度2024 CI={ci_2024:.4f}, 2030 CI={ci_2030:.4f}")
    print(f"  降幅={pct_drop:.2f}% (目标7.8%)")
    print(f"  {'✅' if abs(pct_drop-7.8)<0.5 else '⚠️'} 偏差={pct_drop-7.8:.2f}%")

ci_2035=ws_mod.cell(63,YC+11).value
if ci_2035 and ci_2024:
    pct_drop35=(ci_2024-ci_2035)/ci_2024*100
    print(f"  适度2035 CI={ci_2035:.4f}, 降幅={pct_drop35:.2f}% (目标17.6%)")
    print(f"  {'✅' if abs(pct_drop35-17.6)<0.5 else '⚠️'} 偏差={pct_drop35-17.6:.2f}%")

wb.close()
print("\n完成！")
