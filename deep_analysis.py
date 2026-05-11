#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深度分析Excel模型结构 - 找出每个Step的具体行位置和当前参数
"""

import openpyxl
from openpyxl.utils import get_column_letter

file_path = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)  # 读取计算值

# 详细扫描积极情景Sheet
scenario_name = "积极情景"
ws = wb[scenario_name]

print(f"{'='*120}")
print(f"【{scenario_name}】Sheet完整结构扫描")
print(f"{'='*120}")
print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")

# 找到年份行
year_row = None
for row in range(1, 10):
    for col in range(1, 45):
        if ws.cell(row, col).value == 2024:
            year_row = row
            year_col_start = col
            break
    if year_row:
        break

print(f"年份行: Row {year_row}, 起始列: Col {year_col_start} ({get_column_letter(year_col_start)})")

# 扫描全部65行,显示每行的B列标签和关键年份数据
years_to_check = [2024, 2030, 2035, 2040, 2050, 2060]
year_cols = {}
for year in years_to_check:
    for col in range(year_col_start, year_col_start + 40):
        if ws.cell(year_row, col).value == year:
            year_cols[year] = col

print(f"\n关键年份列位置: {year_cols}")
print(f"\n逐行扫描 (B列标签 | 2024值 | 2030值 | 2035值 | 2060值):")
print("-" * 120)

for row in range(1, ws.max_row + 1):
    b_val = ws.cell(row, 2).value
    if b_val is None:
        continue
    
    b_str = str(b_val).strip()[:55]
    
    # 读取关键年份的值
    vals = []
    for year in [2024, 2030, 2035, 2060]:
        col = year_cols.get(year)
        if col:
            v = ws.cell(row, col).value
            if isinstance(v, float):
                vals.append(f"{v:.4f}" if abs(v) < 100 else f"{v:.0f}")
            elif isinstance(v, int):
                vals.append(str(v))
            else:
                vals.append(str(v)[:20] if v else "")
        else:
            vals.append("")
    
    print(f"Row {row:3d} | {b_str:55s} | {vals[0]:15s} | {vals[1]:15s} | {vals[2]:15s} | {vals[3]:15s}")

# 现在扫描适度情景和保守情景的Step 4参数（短流程占比和氢冶金占比）
print(f"\n{'='*120}")
print("三个情景Step 4关键参数对比 (短流程占比 | 氢冶金占比)")
print(f"{'='*120}")

for scenario in ['积极情景', '适度情景', '保守情景']:
    ws2 = wb[scenario]
    
    # 找年份行
    yr = None
    yc = None
    for row in range(1, 10):
        for col in range(1, 45):
            if ws2.cell(row, col).value == 2024:
                yr = row
                yc = col
                break
        if yr:
            break
    
    # 找短流程占比和氢冶金占比行
    eaf_row = None
    h2_row = None
    
    for row in range(13, 20):
        cell = ws2.cell(row, 2).value
        if cell:
            if "短流程" in str(cell) and "占比" in str(cell):
                eaf_row = row
            if "氢冶金" in str(cell) and "占比" in str(cell):
                h2_row = row
    
    print(f"\n【{scenario}】")
    if eaf_row:
        print(f"  短流程占比 (Row {eaf_row}):")
        for year in [2024, 2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(eaf_row, col).value
            print(f"    {year}: {v}%")
    else:
        print(f"  未找到短流程占比行")
    
    if h2_row:
        print(f"  氢冶金占比 (Row {h2_row}):")
        for year in [2024, 2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(h2_row, col).value
            print(f"    {year}: {v}%")
    else:
        print(f"  未找到氢冶金占比行")

# 扫描Step 8碳强度输出
print(f"\n{'='*120}")
print("三个情景Step 8碳强度实际输出")
print(f"{'='*120}")

for scenario in ['积极情景', '适度情景', '保守情景']:
    ws2 = wb[scenario]
    
    yr = None
    yc = None
    for row in range(1, 10):
        for col in range(1, 45):
            if ws2.cell(row, col).value == 2024:
                yr = row
                yc = col
                break
        if yr:
            break
    
    # 找碳强度行
    ci_row = None
    for row in range(55, ws2.max_row + 1):
        cell = ws2.cell(row, 2).value
        if cell and "碳强度" in str(cell) and "吨钢" in str(cell):
            ci_row = row
            break
    
    # 找CCUS行
    ccus_row = None
    ccus_pct_row = None
    for row in range(50, ws2.max_row + 1):
        cell = ws2.cell(row, 2).value
        if cell:
            if "CCUS" in str(cell) and "捕集" in str(cell):
                ccus_row = row
            if "CCUS" in str(cell) and ("比例" in str(cell) or "占比" in str(cell)):
                ccus_pct_row = row
    
    # 找降碳潜力为负数的行
    neg_rows = []
    for row in range(39, ws2.max_row + 1):
        cell = ws2.cell(row, 2).value
        if cell and "降碳潜力" in str(cell):
            # 检查是否有负值
            for year in [2025, 2030, 2035]:
                col = yc + (year - 2024)
                v = ws2.cell(row, col).value
                if isinstance(v, (int, float)) and v < 0:
                    neg_rows.append((row, cell, year, v))
    
    print(f"\n【{scenario}】")
    if ci_row:
        print(f"  碳强度输出 (Row {ci_row}):")
        for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(ci_row, col).value
            print(f"    {year}: {v}")
    else:
        print(f"  未找到碳强度行")
    
    if ccus_row:
        print(f"  CCUS捕集量 (Row {ccus_row}):")
        for year in [2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(ccus_row, col).value
            print(f"    {year}: {v} 万吨")
    
    if ccus_pct_row:
        print(f"  CCUS比例 (Row {ccus_pct_row}):")
        for year in [2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(ccus_pct_row, col).value
            print(f"    {year}: {v}")
    
    if neg_rows:
        print(f"  ⚠️ 发现降碳潜力为负数的行:")
        for row, label, year, val in neg_rows:
            print(f"    Row {row} '{str(label)[:40]}' - {year}: {val}")
    else:
        print(f"  降碳潜力无负值 ✓")

# 扫描绿电占比
print(f"\n{'='*120}")
print("三个情景绿电占比对比")
print(f"{'='*120}")

for scenario in ['积极情景', '适度情景', '保守情景']:
    ws2 = wb[scenario]
    
    yr = None
    yc = None
    for row in range(1, 10):
        for col in range(1, 45):
            if ws2.cell(row, col).value == 2024:
                yr = row
                yc = col
                break
        if yr:
            break
    
    green_row = None
    for row in range(30, 35):
        cell = ws2.cell(row, 2).value
        if cell and "绿电" in str(cell) and "占比" in str(cell):
            green_row = row
    
    print(f"【{scenario}】绿电占比 (Row {green_row}):")
    if green_row:
        for year in [2024, 2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(green_row, col).value
            print(f"  {year}: {v}%")

# 扫描低碳技术累计降碳潜力
print(f"\n{'='*120}")
print("三个情景低碳技术累计降碳潜力")
print(f"{'='*120}")

for scenario in ['积极情景', '适度情景', '保守情景']:
    ws2 = wb[scenario]
    
    yr = None
    yc = None
    for row in range(1, 10):
        for col in range(1, 45):
            if ws2.cell(row, col).value == 2024:
                yr = row
                yc = col
                break
        if yr:
            break
    
    tech_row = None
    for row in range(34, 40):
        cell = ws2.cell(row, 2).value
        if cell and "累计" in str(cell) and "降碳潜力" in str(cell):
            tech_row = row
    
    print(f"【{scenario}】累计技术降碳潜力 (Row {tech_row}):")
    if tech_row:
        for year in [2024, 2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws2.cell(tech_row, col).value
            print(f"  {year}: {v} tCO₂/t")

wb.close()
print(f"\n{'='*120}")
print("深度分析完成")