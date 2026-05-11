#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
精确复现Excel V11的Step 8计算逻辑，然后调整参数实现目标对齐
"""
import openpyxl
from openpyxl.utils import get_column_letter

file = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'

# 读公式
wb_f = openpyxl.load_workbook(file, data_only=False)
# 读计算值  
wb_v = openpyxl.load_workbook(file, data_only=True)

YEARS = list(range(2024, 2061))

for sc in ['适度情景']:
    ws_f = wb_f[sc]
    ws_v = wb_v[sc]
    
    # 年份列
    yc = 3  # C列=2024
    
    print(f"【{sc}】关键行公式和值对照：")
    print(f"\n--- Row 37 (累计技术 = 新增累计) ---")
    for yr in [2024,2025,2026,2030,2035]:
        col = yc + (yr - 2024)
        f = ws_f.cell(37, col).value
        v = ws_v.cell(37, col).value
        print(f"  {yr} ({get_column_letter(col)}): formula={str(f)[:60]}, value={v}")
    
    print(f"\n--- Row 38 (累计技术+0.066) ---")
    for yr in [2024,2025,2026,2030,2035]:
        col = yc + (yr - 2024)
        f = ws_f.cell(38, col).value
        v = ws_v.cell(38, col).value
        print(f"  {yr}: formula={str(f)[:60]}, value={v}")
    
    print(f"\n--- Row 36 (逐年新增) ---")
    for yr in [2024,2025,2026,2027,2028,2029,2030,2035]:
        col = yc + (yr - 2024)
        f = ws_f.cell(36, col).value
        v = ws_v.cell(36, col).value
        print(f"  {yr}: formula={str(f)[:60]}, value={v}")
    
    print(f"\n--- Row 56 (低碳工艺降碳潜力) ---")
    for yr in [2024,2025,2026,2030,2035]:
        col = yc + (yr - 2024)
        f = ws_f.cell(56, col).value
        v = ws_v.cell(56, col).value
        print(f"  {yr}: formula={str(f)[:60]}, value={v}")

    print(f"\n--- Row 57 (低碳工艺碳排放曲线 = curve_tech) ---")
    for yr in [2024,2025,2026,2030,2035,2040,2050,2060]:
        col = yc + (yr - 2024)
        f = ws_f.cell(57, col).value
        v = ws_v.cell(57, col).value
        print(f"  {yr}: formula={str(f)[:60]}, value={v}")
    
    print(f"\n--- Row 59 (CCUS动态比例) ---")
    for yr in [2024,2025,2026,2030,2035,2040,2050,2060]:
        col = yc + (yr - 2024)
        f = ws_f.cell(59, col).value
        v = ws_v.cell(59, col).value
        print(f"  {yr}: formula={str(f)[:60]}, value={v}")
    
    print(f"\n--- Row 63 (吨钢碳排放强度) ---")
    for yr in [2024,2025,2026,2030,2035,2040,2050,2060]:
        col = yc + (yr - 2024)
        f = ws_f.cell(63, col).value
        v = ws_v.cell(63, col).value
        print(f"  {yr}: formula={str(f)[:60]}, value={v}")

    # 完整瀑布流各步骤2030年的值
    print(f"\n--- 2030年瀑布流详细 ---")
    rows = [41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63]
    col30 = yc + 6  # 2030
    for row in rows:
        a = ws_v.cell(row, 1).value
        v = ws_v.cell(row, col30).value
        if a or v:
            print(f"  Row {row:3d}: {str(a or '')[:50]:50s} = {v}")

wb_f.close()
wb_v.close()
print("\n分析完成")
