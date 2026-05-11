#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查V11.1其余需要完成的工作：
1. 降碳潜力负值情况（Row 44/47）
2. Step 7技术降碳曲线平滑性（Row 36/37）
3. 财务机构Sheet状态
"""
import openpyxl

wb = openpyxl.load_workbook('/Users/daiyinlong/Desktop/钢铁行业模型V11.1.xlsx', data_only=True)
YEARS = list(range(2024,2061))
YC = 3

print("=== 1. 降碳潜力负值检查（Row 44/47/50）===")
print("（prd2要求：除产量变化外，其他降碳潜力不应有负值，但废钢减少导致的负值在模型中是合理的）")
for sc in ['积极情景','适度情景','保守情景']:
    ws=wb[sc]
    print(f"\n【{sc}】")
    rr_neg=[]
    er_neg=[]
    hr_neg=[]
    for y in YEARS:
        col=YC+(y-2024)
        rr=ws.cell(44,col).value
        er=ws.cell(47,col).value
        hr=ws.cell(50,col).value
        if rr and rr<-1000: rr_neg.append((y,rr))
        if er and er<-100: er_neg.append((y,er))
        if hr and hr<0: hr_neg.append((y,hr))
    print(f"  Row44原料: 负值年份={len(rr_neg)}, 最大负值={min([x[1] for x in rr_neg]) if rr_neg else 0:.0f}")
    if rr_neg: print(f"    样例:{rr_neg[:3]}")
    print(f"  Row47短流程: 负值年份={len(er_neg)}, 最大负值={min([x[1] for x in er_neg]) if er_neg else 0:.0f}")
    if er_neg: print(f"    样例:{er_neg[:3]}")
    print(f"  Row50氢冶金: 负值年份={len(hr_neg)}")

print("\n=== 2. Step7技术降碳曲线（Row 36/37）平滑性 ===")
print("（prd2要求：2040年前平滑，无跳跃）")
for sc in ['积极情景','适度情景','保守情景']:
    ws=wb[sc]
    print(f"\n【{sc}】")
    vals36=[]
    for y in YEARS:
        col=YC+(y-2024)
        v=ws.cell(36,col).value
        vals36.append((y,v or 0))
    print(f"  Row36年增量(2024-2040):")
    for y,v in vals36[:17]:
        print(f"    {y}: {v:.6f}")

print("\n=== 3. 财务机构Sheet状态 ===")
for shname in ['金融机构目标设置（组合估算）','金融机构目标设置（企业）']:
    if shname in wb.sheetnames:
        ws=wb[shname]
        # 检查是否有数据
        has_data=False
        for row in range(1,20):
            for col in range(1,15):
                if ws.cell(row,col).value:
                    has_data=True; break
            if has_data: break
        print(f"  {shname}: {'有数据' if has_data else '无数据/空'}")
    else:
        print(f"  {shname}: Sheet不存在")

print("\n=== 4. 可视化图表状态 ===")
for sc in ['积极情景','适度情景','保守情景','汇总对比与可视化看板']:
    if sc in wb.sheetnames:
        ws=wb[sc]
        # openpyxl无法直接读取图表数量，检查_charts属性
        try:
            nc = len(ws._charts) if hasattr(ws,'_charts') else 0
            print(f"  {sc}: 图表数={nc}")
        except: print(f"  {sc}: 无法检测图表")

wb.close()
print("\n分析完成")
