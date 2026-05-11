#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
钢铁行业减碳模型分析脚本 V2 - 详细扫描Excel结构
"""

import openpyxl
from openpyxl.utils import get_column_letter

# 读取Excel文件
file_path = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)  # 读取计算后的值

print("=" * 100)
print("钢铁行业减碳模型V11详细结构分析")
print("=" * 100)

# 分析积极情景Sheet
scenario_name = "积极情景"
if scenario_name in wb.sheetnames:
    ws = wb[scenario_name]
    
    print(f"\n【{scenario_name}】Sheet结构扫描:")
    print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")
    
    # 扫描前70行,显示B列和C列的内容
    print("\n前70行关键内容 (B列标签 | C列数据):")
    print("-" * 100)
    
    for row in range(1, min(71, ws.max_row + 1)):
        b_val = ws.cell(row, 2).value  # B列
        c_val = ws.cell(row, 3).value  # C列
        
        # 如果B列有内容,显示这一行
        if b_val:
            b_str = str(b_val)[:50]  # 截断长文本
            c_str = str(c_val)[:30] if c_val is not None else "None"
            print(f"Row {row:2d} | {b_str:50s} | {c_str}")
    
    # 找到年份行
    print("\n\n查找年份行:")
    for row in range(1, 20):
        for col in range(1, 10):
            val = ws.cell(row, col).value
            if val == 2024:
                print(f"找到2024年: Row {row}, Col {get_column_letter(col)}")
                
                # 显示该行的数据
                print("\n年份行数据:")
                year_values = []
                for c in range(col, min(col + 40, ws.max_column + 1)):
                    year_val = ws.cell(row, c).value
                    if year_val and isinstance(year_val, (int, float)) and 2024 <= year_val <= 2060:
                        year_values.append((c, year_val))
                
                for c, y in year_values[:10]:  # 显示前10个
                    print(f"  Col {get_column_letter(c)}: {y}")
                
                # 扫描该年份行下方的数据
                print(f"\n从年份行 (Row {row}) 往下30行的数据扫描:")
                print("-" * 100)
                for r in range(row + 1, min(row + 31, ws.max_row + 1)):
                    b_val = ws.cell(r, 2).value
                    # 显示2024年(第一个数据列)的值
                    data_val = ws.cell(r, col).value
                    
                    if b_val:
                        b_str = str(b_val)[:60]
                        d_str = str(data_val)[:30] if data_val is not None else "None"
                        print(f"Row {r:2d} | {b_str:60s} | 2024值: {d_str}")
                
                break
        else:
            continue
        break

# 分析Dashboard
print("\n\n" + "=" * 100)
print("【汇总对比与可视化看板】Sheet结构扫描")
print("=" * 100)

dashboard_name = "汇总对比与可视化看板"
if dashboard_name in wb.sheetnames:
    ws_d = wb[dashboard_name]
    print(f"总行数: {ws_d.max_row}, 总列数: {ws_d.max_column}")
    
    # 扫描前50行
    print("\n前50行数据扫描 (A-F列):")
    print("-" * 100)
    print(f"{'Row':4s} | {'A列':15s} | {'B列(保守)':15s} | {'C列':10s} | {'D列(适度)':15s} | {'E列':10s} | {'F列(积极)':15s}")
    print("-" * 100)
    
    for row in range(1, min(51, ws_d.max_row + 1)):
        a_val = ws_d.cell(row, 1).value
        b_val = ws_d.cell(row, 2).value
        c_val = ws_d.cell(row, 3).value
        d_val = ws_d.cell(row, 4).value
        e_val = ws_d.cell(row, 5).value
        f_val = ws_d.cell(row, 6).value
        
        # 如果A列有数据,或者B/D/F列有数值数据
        if a_val or (isinstance(b_val, (int, float)) and b_val > 0):
            a_str = str(a_val)[:15] if a_val else ""
            b_str = f"{b_val:.4f}" if isinstance(b_val, (int, float)) else str(b_val)[:15] if b_val else ""
            c_str = str(c_val)[:10] if c_val else ""
            d_str = f"{d_val:.4f}" if isinstance(d_val, (int, float)) else str(d_val)[:15] if d_val else ""
            e_str = str(e_val)[:10] if e_val else ""
            f_str = f"{f_val:.4f}" if isinstance(f_val, (int, float)) else str(f_val)[:15] if f_val else ""
            
            print(f"{row:4d} | {a_str:15s} | {b_str:15s} | {c_str:10s} | {d_str:15s} | {e_str:10s} | {f_str:15s}")

wb.close()

print("\n" + "=" * 100)
print("分析完成")
print("=" * 100)
