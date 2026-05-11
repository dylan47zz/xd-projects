#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
扫描适度情景完整参数
"""

import openpyxl

file_path = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

scenario = "适度情景"
ws = wb[scenario]

yr = None
yc = None
for row in range(1, 10):
    for col in range(1, 45):
        if ws.cell(row, col).value == 2024:
            yr = row
            yc = col
            break
    if yr:
        break

print(f"【{scenario}】完整参数扫描")
print(f"年份行: Row {yr}, 起始列: Col {yc}")

# Step 3 粗钢产量
print(f"\nStep 3 粗钢产量 (Row 13):")
for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
    col = yc + (year - 2024)
    v = ws.cell(13, col).value
    print(f"  {year}: {v} 万吨 ({v/10000:.2f}亿吨)")

# Step 4 流程结构
print(f"\nStep 4 流程结构 (Row 15-16):")
for row in [15, 16]:
    label = ws.cell(row, 1).value or ws.cell(row, 2).value
    vals = ""
    for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
        col = yc + (year - 2024)
        v = ws.cell(row, col).value
        vals += f" | {year}:{v}"
    print(f"  Row {row}: {str(label)[:40]}{vals}")

# Step 5 废钢资源量参考
print(f"\nStep 5 废钢资源量参考 (Row 22):")
for year in [2024, 2030, 2035, 2040, 2050, 2060]:
    col = yc + (year - 2024)
    v = ws.cell(22, col).value
    print(f"  {year}: {v}万吨")

# 废钢消耗总量
print(f"\n  废钢消耗总量 (Row 28):")
for year in [2024, 2030, 2035, 2040, 2050, 2060]:
    col = yc + (year - 2024)
    v = ws.cell(28, col).value
    print(f"  {year}: {v}万吨")

# Step 6 绿电占比
print(f"\nStep 6 绿电占比 (Row 32):")
for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
    col = yc + (year - 2024)
    v = ws.cell(32, col).value
    print(f"  {year}: {v}")

# Step 7 低碳技术
print(f"\nStep 7 低碳技术 (Row 36-38):")
for row in [36, 37, 38]:
    b_val = ws.cell(row, 2).value
    vals = ""
    for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
        col = yc + (year - 2024)
        v = ws.cell(row, col).value
        if v is not None:
            vals += f" | {year}:{v}"
    print(f"  Row {row}: {str(b_val)[:40]}{vals}")

# Step 8 碳强度 vs Dashboard
dashboard_target = {
    2024: 1.9500, 2025: 1.9540, 2030: 1.7980, 2035: 1.6070,
    2040: 1.1500, 2050: 0.4000, 2060: 0.0500
}

print(f"\nStep 8 碳强度 (Row 63) vs Dashboard目标:")
for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
    col = yc + (year - 2024)
    actual = ws.cell(63, col).value
    target = dashboard_target[year]
    if actual:
        deviation = actual - target
        pct_dev = deviation / target * 100
        # 计算较2024年的降幅
        vs_2024 = (actual - 1.95) / 1.95 * 100
        print(f"  {year}: 实际={actual:.4f}, 目标={target:.4f}, 偏差={deviation:.4f} ({pct_dev:+.1f}%), 较2024年变化={vs_2024:+.1f}%")

# 适度目标2030年较2024年下降7.8%
print(f"\n  适度目标验证: 2030年较2024年降幅 = (1.798-1.95)/1.95 = {(1.798-1.95)/1.95*100:.1f}%")
print(f"  适度目标验证: 2035年较2024年降幅 = (1.607-1.95)/1.95 = {(1.607-1.95)/1.95*100:.1f}%")

# CCUS
print(f"\nCCUS (Row 59-60):")
for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
    col = yc + (year - 2024)
    pct = ws.cell(59, col).value
    amount = ws.cell(60, col).value
    print(f"  {year}: 比例={pct}, 捕集量={amount}万吨")

# 降碳潜力负数检查
print(f"\n降碳潜力负数检查:")
for row in [42, 44, 47, 50, 53, 56]:
    a_val = ws.cell(row, 1).value
    label = str(a_val)[:50]
    neg_years = []
    for year in range(2025, 2061):
        col = yc + (year - 2024)
        v = ws.cell(row, col).value
        if isinstance(v, (int, float)) and v < 0:
            neg_years.append((year, v))
    if neg_years:
        print(f"  ⚠️ Row {row} ({label}): 有负值! {neg_years[:3]}...")
    else:
        print(f"  ✓ Row {row} ({label}): 无负值")

wb.close()
print("\n分析完成")