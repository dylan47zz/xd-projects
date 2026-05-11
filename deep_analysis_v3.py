#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
精确扫描Excel模型 - 逐行逐列分析
"""

import openpyxl
from openpyxl.utils import get_column_letter

file_path = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

# 扫描积极情景的全部行，展示每一列的内容
scenario = "积极情景"
ws = wb[scenario]

print(f"{'='*150}")
print(f"【{scenario}】逐行逐列详细扫描")
print(f"{'='*150}")

# 找年份行
yr = None
yc = None
for row in range(1, 10):
    for col in range(1, 45):
        if ws.cell(row, col).value == 2024:
            yr = row
            yc = col
            break
    if yr:
        break

print(f"年份行: Row {yr}, 起始列: Col {yc}")
print(f"\n年份序列 (Row {yr}):")
for col in range(yc, yc + 37):
    v = ws.cell(yr, col).value
    if v:
        print(f"  Col {get_column_letter(col)} ({col}): {v}")

# 扫描每一行的所有列内容
print(f"\n逐行详细内容:")
print("-" * 150)

for row in range(1, ws.max_row + 1):
    # 先扫描A-B列
    a_val = ws.cell(row, 1).value
    b_val = ws.cell(row, 2).value
    
    # 扫描所有列，找有内容的
    row_content = []
    for col in range(1, 44):
        v = ws.cell(row, col).value
        if v is not None and v != '' and col != yr:  # 跳过空值和年份行
            # 不扫描年份列区域的数据列
            if col >= yc and col < yc + 37 and row != yr:
                continue  # 数据区域的内容单独处理
            row_content.append((col, v))
    
    if row_content or b_val:
        line = f"Row {row:3d}: "
        for col, v in row_content[:5]:
            col_letter = get_column_letter(col)
            v_str = str(v)[:40]
            line += f"| {col_letter}:{v_str} "
        
        # 现在显示数据区域的关键年份数据
        if row > yr:
            for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
                col = yc + (year - 2024)
                v = ws.cell(row, col).value
                if v is not None:
                    if isinstance(v, float):
                        if abs(v) < 1:
                            v_str = f"{v:.4f}"
                        elif abs(v) < 100:
                            v_str = f"{v:.2f}"
                        else:
                            v_str = f"{v:.0f}"
                    elif isinstance(v, int):
                        v_str = str(v)
                    else:
                        v_str = str(v)[:30]
                    line += f"| {year}:{v_str} "
        
        print(line)

# 也读取公式版本的几个关键行
print(f"\n{'='*150}")
print("读取关键行的公式 (data_only=False)")
print(f"{'='*150}")

wb_formula = openpyxl.load_workbook(file_path, data_only=False)
ws_f = wb_formula[scenario]

for row in [12, 14, 15, 30, 31, 32, 34, 37, 38, 39, 41, 42, 43, 44, 45, 48, 51, 53, 54, 56, 57, 59, 60, 62, 63]:
    b_val = ws_f.cell(row, 2).value
    if b_val:
        print(f"\nRow {row}: B='{str(b_val)[:50]}'")
        # 显示2024年的公式
        v = ws_f.cell(row, yc).value
        if v:
            print(f"  2024 公式/值: {str(v)[:80]}")
        # 显示2030年的公式
        v = ws_f.cell(row, yc + 6).value
        if v:
            print(f"  2030 公式/值: {str(v)[:80]}")

wb_formula.close()
wb.close()
print("\n分析完成")