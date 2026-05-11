#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""精确复现V11公式，找出curve_tech偏低的真正原因"""

import openpyxl

f = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(f, data_only=True)
ws = wb['适度情景']

YC = 3  # C=2024

print("=== V11 适度情景 2025-2035年 完整数据 ===")
for yr in [2025, 2026, 2027, 2028, 2029, 2030]:
    col = YC + (yr - 2024)
    print(f"\n{yr}年:")
    print(f"  Row13 产量: {ws.cell(13, col).value}")
    print(f"  Row15 短流程%: {ws.cell(15, col).value}")
    print(f"  Row16 氢冶金%: {ws.cell(16, col).value}")
    print(f"  Row22 废钢资源量参考: {ws.cell(22, col).value}")
    print(f"  Row25 短流程废钢消耗: {ws.cell(25, col).value}")
    print(f"  Row28 废钢消耗总量: {ws.cell(28, col).value}")
    print(f"  Row32 绿电占比: {ws.cell(32, col).value}")
    print(f"  Row36 逐年新增技术: {ws.cell(36, col).value}")
    print(f"  Row37 累计新增: {ws.cell(37, col).value}")
    print(f"  Row44 原料降碳潜力: {ws.cell(44, col).value}")
    print(f"  Row47 短流程降碳潜力: {ws.cell(47, col).value}")
    print(f"  Row57 curve_tech: {ws.cell(57, col).value}")
    print(f"  Row59 CCUS比例: {ws.cell(59, col).value}")
    print(f"  Row63 CI: {ws.cell(63, col).value}")

print("\n\n=== V11 三情景 基准年2024年废钢消耗 ===")
for sc in ['积极情景', '适度情景', '保守情景']:
    ws2 = wb[sc]
    print(f"\n{sc}:")
    print(f"  Row25 短流程废钢消耗: {ws2.cell(25, 3).value}")
    print(f"  Row28 废钢消耗总量: {ws2.cell(28, 3).value}")
    print(f"  Row22 废钢资源量参考: {ws2.cell(22, 3).value}")

wb.close()

# 关键验证：用V11原始参数复现2030年
print("\n\n=== Python复现V11适度情景2025年 ===")
P = 96081
eaf = 0.10; h2 = 0.011; bf = 1-eaf-h2
P_bf,P_eaf,P_h2 = P*bf,P*eaf,P*h2
print(f"产量: {P}, EAF={eaf*100}%, H2={h2*100}%, BF={bf*100:.1f}%")
print(f"P_bf={P_bf:.1f}, P_eaf={P_eaf:.1f}, P_h2={P_h2:.1f}")

scrap_bf = P_bf*0.30
scrap_eaf = P_eaf*1.10
scrap_h2  = P_h2*0.50
scrap_tot = scrap_bf+scrap_eaf+scrap_h2
print(f"scrap_bf={scrap_bf:.1f}, scrap_eaf={scrap_eaf:.1f}, scrap_h2={scrap_h2:.1f}")
print(f"scrap_tot={scrap_tot:.1f} (基准38394.44)")

raw_reduce = (scrap_tot-38394.44)/1.1*1.35
eaf_reduce = (scrap_eaf-11055.99)*1.26
print(f"raw_reduce={raw_reduce:.1f} (V11实际约-2052)")
print(f"eaf_reduce={eaf_reduce:.1f} (V11实际约-614)")

# 说明：废钢消耗总量由流程结构决定，2025年废钢消耗<基准年38394，导致原料降碳为负
# 这是合理的（产量下降，废钢消耗也下降）
