#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
扫描三个情景的完整参数和Step 8输出对比Dashboard目标
"""

import openpyxl
from openpyxl.utils import get_column_letter

file_path = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

# Dashboard目标值
dashboard_targets = {
    '积极情景': {
        2024: 1.9500, 2025: 1.9470, 2030: 1.7440, 2035: 1.3700,
        2040: 0.8500, 2050: 0.1700, 2060: 0.0100
    },
    '适度情景': {
        2024: 1.9500, 2025: 1.9540, 2030: 1.7980, 2035: 1.6070,
        2040: 1.1500, 2050: 0.4000, 2060: 0.0500
    },
    '保守情景': {
        2024: 1.9500, 2025: 1.9580, 2030: 1.8740, 2035: 1.7510,
        2040: 1.4500, 2050: 0.7500, 2060: 0.1200
    }
}

for scenario in ['积极情景', '适度情景', '保守情景']:
    ws = wb[scenario]
    
    # 找年份行
    yr = None
    yc = None
    for row in range(1, 10):
        for col in range(1, 45):
            if ws.cell(row, col).value == 2024:
                yr = row
                yc = col
                break
        if yr:
            break
    
    print(f"\n{'='*120}")
    print(f"【{scenario}】关键参数和偏差分析")
    print(f"{'='*120}")
    
    # Step 3: 粗钢产量 (Row 13)
    print(f"\nStep 3 粗钢产量 (Row 13):")
    for year in [2024, 2030, 2035, 2040, 2050, 2060]:
        col = yc + (year - 2024)
        v = ws.cell(13, col).value
        print(f"  {year}: {v} 万吨 ({v/10000:.2f}亿吨)")
    
    # Step 4: 流程结构 (Row 14-20)
    print(f"\nStep 4 流程结构 (Row 14-20):")
    for row in range(14, 21):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        if a_val or b_val:
            label = str(a_val or b_val)[:50]
            vals = ""
            for year in [2024, 2030, 2035, 2040, 2050, 2060]:
                col = yc + (year - 2024)
                v = ws.cell(row, col).value
                if v is not None:
                    vals += f" | {year}:{v}"
            print(f"  Row {row}: {label}{vals}")
    
    # Step 5: 铁素资源 (Row 21-29)
    print(f"\nStep 5 铁素资源 (Row 21-29):")
    for row in range(21, 30):
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        if a_val or b_val:
            label = str(a_val or b_val)[:50]
            vals = ""
            for year in [2024, 2030, 2035, 2060]:
                col = yc + (year - 2024)
                v = ws.cell(row, col).value
                if v is not None:
                    vals += f" | {year}:{v}"
            print(f"  Row {row}: {label}{vals}")
    
    # Step 6: 绿电占比 (Row 32)
    print(f"\nStep 6 绿电占比 (Row 32):")
    for year in [2024, 2030, 2035, 2040, 2050, 2060]:
        col = yc + (year - 2024)
        v = ws.cell(32, col).value
        print(f"  {year}: {v}")
    
    # Step 7: 低碳技术 (Row 36-38)
    print(f"\nStep 7 低碳技术 (Row 36-38):")
    for row in range(36, 39):
        b_val = ws.cell(row, 2).value
        vals = ""
        for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
            col = yc + (year - 2024)
            v = ws.cell(row, col).value
            if v is not None:
                vals += f" | {year}:{v}"
        print(f"  Row {row}: {str(b_val)[:40]}{vals}")
    
    # Step 8: 碳强度输出 vs Dashboard目标
    print(f"\nStep 8 碳强度 (Row 63) vs Dashboard目标:")
    target = dashboard_targets[scenario]
    
    for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
        col = yc + (year - 2024)
        actual = ws.cell(63, col).value
        target_val = target[year]
        if actual and target_val:
            deviation = actual - target_val
            pct_dev = deviation / target_val * 100
            print(f"  {year}: 实际={actual:.4f}, 目标={target_val:.4f}, 偏差={deviation:.4f} ({pct_dev:+.1f}%)")
    
    # CCUS配平比例和捕集量 (Row 59-60)
    print(f"\nCCUS (Row 59-60):")
    for year in [2024, 2025, 2030, 2035, 2040, 2050, 2060]:
        col = yc + (year - 2024)
        pct = ws.cell(59, col).value
        amount = ws.cell(60, col).value
        print(f"  {year}: 比例={pct}, 捕集量={amount}万吨")
    
    # 检查降碳潜力是否有负数 (Row 42,44,47,50,53,56)
    print(f"\n降碳潜力行检查 (是否有负数):")
    for row in [42, 44, 47, 50, 53, 56]:
        a_val = ws.cell(row, 1).value
        b_val = ws.cell(row, 2).value
        label = str(a_val or b_val)[:50]
        neg_years = []
        for year in range(2025, 2061):
            col = yc + (year - 2024)
            v = ws.cell(row, col).value
            if isinstance(v, (int, float)) and v < 0:
                neg_years.append((year, v))
        if neg_years:
            print(f"  ⚠️ Row {row} ({label}): 有负值! {neg_years[:5]}...")
        else:
            print(f"  ✓ Row {row} ({label}): 无负值")

wb.close()
print("\n分析完成")