#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将Row 59 CCUS比例从固定数值改为Excel公式：
=MAX(0, MIN(0.98, IF(I57>0, 1-'汇总对比与可视化看板'!F9*I13/I57, 0)))
使CCUS动态响应其他参数变化

Dashboard位置：
- 年份2024在Row 3, 2060在Row 39
- 积极情景CI: Col F (F3=1.95, F4=1.947, ..., F39=0.01)
- 适度情景CI: Col D (D3=1.95, D4=1.954, ..., D39=0.05)
- 保守情景CI: Col B (B3=1.95, B4=1.958, ..., B39=0.12)
"""
import openpyxl
from openpyxl.utils import get_column_letter

OUTPUT = '/Users/daiyinlong/Desktop/钢铁行业模型V11.1.xlsx'
DASH = '汇总对比与可视化看板'
YEARS = list(range(2024, 2061))
YC = 3

# Dashboard目标CI列
DASH_COL = {'积极情景': 'F', '适度情景': 'D', '保守情景': 'B'}
DASH_ROW_START = 3  # 2024年在第3行

# 关闭Excel确保文件可写
import subprocess
subprocess.run(['osascript', '-e', 
    'tell application "Microsoft Excel" to close every workbook without saving'], 
    capture_output=True)

import time; time.sleep(1)

wb = openpyxl.load_workbook(OUTPUT)

for sc in ['积极情景', '适度情景', '保守情景']:
    ws = wb[sc]
    dc = DASH_COL[sc]
    
    for y in YEARS:
        col = YC + (y - 2024)
        cl  = get_column_letter(col)
        # Dashboard中的行号 = DASH_ROW_START + (y - 2024)
        dash_row = DASH_ROW_START + (y - 2024)
        
        if y == 2024:
            # 2024年CCUS=0（基准年）
            ws.cell(59, col).value = 0
        else:
            # 动态公式：从Dashboard引用目标CI
            # 公式：=MAX(0, MIN(0.98, IF({cl}57>0, 1-'{DASH}'!{dc}{dr}*{cl}13/{cl}57, 0)))
            formula = (f"=MAX(0,MIN(0.98,IF({cl}57>0,"
                      f"1-'{DASH}'!{dc}{dash_row}*{cl}13/{cl}57,0)))")
            ws.cell(59, col).value = formula
    
    print(f"✅ {sc} Row59 改为动态公式")

wb.save(OUTPUT)
wb.close()
print(f"\n已保存: {OUTPUT}")
print("示例公式（适度情景2030年）:")
print(f"  =MAX(0,MIN(0.98,IF(I57>0,1-'{DASH}'!D9*I13/I57,0)))")
