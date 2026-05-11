#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将Row 59 CCUS比例改为公式: =1-目标CI*产量/curve_tech（Row57）
这样CCUS就能动态响应其他参数变化

Dashboard目标CI的位置（汇总对比与可视化看板）：
- 积极情景：列F，第3行开始
- 适度情景：列D，第3行开始  
- 保守情景：列B，第3行开始
年份行从第3行开始，第3行=2024，第4行=2025，...第39行=2060

公式：=MAX(0,MIN(0.98,IF(I57>0,1-'汇总对比与可视化看板'!F9*I13/I57,0)))
其中I列对应2030年（年份偏移计算）
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os, shutil

# 验证Dashboard中目标CI的位置
SOURCE = '/Users/daiyinlong/main/xd-projects/副本钢铁行业减碳基准路径模型_V11.xlsx'
OUTPUT = os.path.expanduser('~/Desktop/钢铁行业模型V11.1.xlsx')

wb_src = openpyxl.load_workbook(SOURCE, data_only=True)
dash = wb_src['汇总对比与可视化看板']

print("=== 验证Dashboard目标CI位置 ===")
# 先确定年份行
print("Col A,B,C,D,E,F 前40行：")
for r in range(1, 45):
    vals = []
    for c in range(1, 7):
        v = dash.cell(r, c).value
        if v is not None:
            vals.append(f"({r},{c})={v}")
    if vals:
        print(f"  {vals}")

wb_src.close()
